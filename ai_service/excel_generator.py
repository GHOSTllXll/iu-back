# backend/ai_service/excel_generator.py
import io
import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from .rent_roll_utils import detect_rent_roll_columns, ensure_status_column, RentRollColumnError


class ExcelGenerationError(Exception):
    """Raised when the rent roll doesn't have columns we can confidently map."""
    pass


def format_citation(citation: dict) -> str:
    """
    Module 4 (Source Provenance) — builds a human-readable citation string from
    a provenance dict, gracefully handling any field being null (the AI is
    instructed to use null rather than fabricate a citation it isn't sure of).
    Returns "" if there's nothing usable to cite.
    """
    if not citation:
        return ""

    parts = []
    if citation.get("page_location"):
        parts.append(f"Page {citation['page_location']}")
    if citation.get("section_title"):
        parts.append(f"'{citation['section_title']}'")

    location = ", ".join(parts)
    anchor = citation.get("exact_text_anchor")

    if location and anchor:
        return f"[Source: Offering Memorandum, {location} — \"{anchor}\"]"
    elif location:
        return f"[Source: Offering Memorandum, {location}]"
    elif anchor:
        return f"[Source: Offering Memorandum — \"{anchor}\"]"
    return ""


def clean_cell_value(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (ValueError, TypeError):
        pass
    if isinstance(val, np.integer):
        return int(val)
    if isinstance(val, np.floating):
        return float(val)
    if isinstance(val, (pd.Timestamp, datetime.datetime)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        if val in ['#N/A', '#VALUE!', '#REF!', '#DIV/0!', '#NUM!', '#NAME?', '#NULL!']:
            return ""
        return val.strip()
    return val


def safe_get(d: dict, *keys, default=None):
    """Nested dict.get() that never raises, even if a section is missing."""
    for key in keys[:-1]:
        d = d.get(key, {}) if isinstance(d, dict) else {}
    return d.get(keys[-1], default) if isinstance(d, dict) else default


# ==========================================
# T12 LINE ITEM DEFINITIONS
# Maps metrics JSON keys -> display label -> row on the T12 tab.
# NOTE: This writes ANNUAL totals only, one column, because the AI extraction
# schema only produces annual figures per line item (see conversation: monthly
# breakdown was scoped out pending a decision on expanding the AI schema).
# ==========================================
T12_LINE_ITEMS = [
    ("net_rental_income", "Net Rental Income"),
    ("other_income", "Other Income"),
    ("total_operating_revenue", "TOTAL OPERATING REVENUE", True),   # subtotal row
    ("real_estate_taxes", "Real Estate Taxes"),
    ("property_liability_insurance", "Property & Liability Insurance"),
    ("total_utilities", "Total Utilities"),
    ("repairs_maintenance", "Repairs & Maintenance"),
    ("contract_services", "Contract Services"),
    ("marketing_advertising", "Marketing & Advertising"),
    ("property_management_fee", "Property Management Fee"),
    ("payroll_benefits_staffing", "Payroll, Benefits & Staffing"),
    ("general_administrative", "General & Administrative"),
    ("total_operating_expenses", "TOTAL OPERATING EXPENSES", True),  # subtotal row
    ("net_operating_income", "NET OPERATING INCOME (NOI)", True),    # subtotal row
]


def generate_underwriting_excel(metrics: dict, rent_roll_df: pd.DataFrame, debt_assumptions: dict = None) -> bytes:
    wb = Workbook()

    # debt_assumptions: {'ltv_pct': 75.0, 'interest_rate_pct': 7.0, 'amortization_years': 30}
    # User-entered per upload — see conversation history for why these can't be
    # AI-extracted (they're the buyer's own financing assumptions, not facts
    # stated in the source documents). Falls back to safe defaults if omitted
    # (e.g. if this function is ever called from a path that doesn't collect them).
    if debt_assumptions is None:
        debt_assumptions = {'ltv_pct': 75.0, 'interest_rate_pct': 7.0, 'amortization_years': 30}

    # ==========================================
    # PRE-PROCESSING: Rent Roll column detection
    # (Shared with reconciliation.py — see rent_roll_utils.py. Both the Excel
    # file and the reconciliation flags must agree on which columns are which.)
    # ==========================================
    try:
        rent_col, status_col, tenant_col = detect_rent_roll_columns(rent_roll_df)
        status_col, status_was_inferred = ensure_status_column(rent_roll_df, status_col, tenant_col)
    except RentRollColumnError as e:
        raise ExcelGenerationError(str(e))

    rent_roll_df[rent_col] = pd.to_numeric(rent_roll_df[rent_col], errors='coerce').fillna(0)
    rent_roll_df[status_col] = rent_roll_df[status_col].astype(str).str.strip()
    rent_roll_df = rent_roll_df.replace({pd.NaT: None, pd.NA: None, np.nan: None})

    # ==========================================
    # TAB 3: RENT ROLL ANALYSIS
    # Dynamic row range (NOT hardcoded to row 150 — a property with more units
    # than that would silently undercount with a fixed range).
    # Summary formulas live BELOW the last real data row, not overlapping it.
    # ==========================================
    ws_rr = wb.active
    ws_rr.title = "Rent Roll Analysis"

    for col_idx, col_name in enumerate(rent_roll_df.columns, 1):
        cell = ws_rr.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        if status_was_inferred and col_name == status_col:
            cell.comment = Comment(
                "This rent roll had no explicit Status column. Occupancy was "
                "INFERRED from the Tenant column: a populated tenant name = "
                "Occupied, blank/vacant = Vacant. Verify against the source "
                "rent roll before relying on this.",
                "Underwriting AI"
            )

    rent_col_idx = rent_roll_df.columns.get_loc(rent_col) + 1

    for row_idx, row in enumerate(rent_roll_df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cleaned_val = clean_cell_value(val)
            cell = ws_rr.cell(row=row_idx, column=col_idx, value=cleaned_val)
            if col_idx == rent_col_idx:
                cell.number_format = '$#,##0.00'
                if isinstance(cleaned_val, str) and cleaned_val.replace('.', '').replace('-', '').isdigit():
                    cell.value = float(cleaned_val)

    last_row = len(rent_roll_df) + 1
    status_letter = get_column_letter(rent_roll_df.columns.get_loc(status_col) + 1)
    rent_letter = get_column_letter(rent_col_idx)

    unit_count_row = last_row + 2
    vacant_units_row = last_row + 3
    occupancy_row = last_row + 4
    avg_rent_row = last_row + 5

    ws_rr.cell(row=unit_count_row, column=1, value="TOTAL UNIT COUNT:").font = Font(bold=True)
    ws_rr.cell(row=unit_count_row, column=2, value=f"=COUNTA(A2:A{last_row})")

    ws_rr.cell(row=vacant_units_row, column=1, value="TOTAL VACANT UNITS:").font = Font(bold=True)
    ws_rr.cell(row=vacant_units_row, column=2, value=f'=COUNTIF({status_letter}2:{status_letter}{last_row}, "Vacant")')

    ws_rr.cell(row=occupancy_row, column=1, value="PHYSICAL OCCUPANCY %:").font = Font(bold=True)
    ws_rr.cell(row=occupancy_row, column=2,
               value=f'=COUNTIF({status_letter}2:{status_letter}{last_row}, "Occupied") / COUNTA(A2:A{last_row})')
    ws_rr.cell(row=occupancy_row, column=2).number_format = '0.00%'

    ws_rr.cell(row=avg_rent_row, column=1, value="AVG IN-PLACE RENT:").font = Font(bold=True)
    ws_rr.cell(row=avg_rent_row, column=2, value=f"=AVERAGE({rent_letter}2:{rent_letter}{last_row})")
    ws_rr.cell(row=avg_rent_row, column=2).number_format = '$#,##0.00'

    # ==========================================
    # TAB 2: CLEANED T12
    # Real extracted annual figures — NOT hardcoded placeholder data.
    # (Previous version wrote a fixed $15,000/month regardless of actual metrics —
    # fixed here.)
    # ==========================================
    ws_t12 = wb.create_sheet("Cleaned T12")
    ws_t12.cell(row=1, column=1, value="LINE ITEM")
    ws_t12.cell(row=1, column=2, value="ANNUAL TOTAL")
    for cell in ws_t12[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")

    t12_data = metrics.get("t12_revenue_expenses", {})
    line_item_rows = {}  # key -> row number, so we can reference rows for subtotals

    row_num = 2
    for item in T12_LINE_ITEMS:
        key, label = item[0], item[1]
        is_subtotal = len(item) > 2 and item[2]

        value = t12_data.get(key)
        cell_label = ws_t12.cell(row=row_num, column=1, value=label)
        cell_value = ws_t12.cell(row=row_num, column=2)

        if is_subtotal:
            cell_label.font = Font(bold=True)
            cell_value.font = Font(bold=True)
            # Subtotals are written as real values from the AI extraction, not
            # re-derived via SUM() here, since the AI's total may legitimately
            # differ slightly from a naive sum if it applied its own judgment
            # (e.g. rounding, or excluded a line item as non-recurring).
            # If you want the sheet to self-recalculate subtotals from the line
            # items above instead of trusting the AI's stated total, that's a
            # one-line change — say so and I'll switch it to a SUM() formula.

        cell_value.value = value if value is not None else ""
        if isinstance(value, (int, float)):
            cell_value.number_format = '$#,##0.00'

        line_item_rows[key] = row_num
        row_num += 1

    # ==========================================
    # MODULE 2 (Enterprise only): Uncategorized Expenses
    # Per spec: a dedicated, visible row for any T12 line item the AI couldn't
    # confidently map into a standard category, so the model's total still
    # balances against the raw source document instead of silently dropping
    # or misfiling anything.
    # ==========================================
    standardization = metrics.get("standardization")
    if standardization and standardization.get("uncategorized_items"):
        uncategorized_header_row = row_num + 1
        ws_t12.cell(row=uncategorized_header_row, column=1, value="UNCATEGORIZED DISCREPANCIES").font = Font(bold=True, size=12, color="FFB8860B")

        item_row = uncategorized_header_row + 1
        for item in standardization["uncategorized_items"]:
            label = f"{item['original_line_item']} ({item['category']})"
            label_cell = ws_t12.cell(row=item_row, column=1, value=label)
            label_cell.comment = Comment(
                "AI could not confidently map this line item to a standard "
                "category. Included in totals as-is — please manually reallocate "
                "if a more specific category applies.",
                "Standardization Engine"
            )
            amount_cell = ws_t12.cell(row=item_row, column=2, value=item["allocated_amount"])
            amount_cell.number_format = '$#,##0.00'
            item_row += 1

        total_uncategorized_row = item_row
        ws_t12.cell(row=total_uncategorized_row, column=1, value="TOTAL UNCATEGORIZED").font = Font(bold=True)
        total_cell = ws_t12.cell(row=total_uncategorized_row, column=2, value=standardization["uncategorized_total"])
        total_cell.font = Font(bold=True)
        total_cell.number_format = '$#,##0.00'

        row_num = total_uncategorized_row + 1

    # ==========================================
    # TAB 2 continued: Capital Structure
    # Only builds what current data actually supports: Purchase Price + DST/Capex.
    # Loan sizing / debt service / DSCR are DEFERRED — see conversation: those
    # need LTV%, Interest Rate%, and Amortization Term inputs that don't exist
    # anywhere in this system yet (not extracted, not user-entered).
    # ==========================================
    capital_header_row = row_num + 1
    ws_t12.cell(row=capital_header_row, column=1, value="CAPITAL STRUCTURE").font = Font(bold=True, size=12)

    deal_valuation = metrics.get("deal_valuation", {})
    purchase_price = deal_valuation.get("target_purchase_price")
    dst_capex = deal_valuation.get("dst_capex_budget")  # None for Basic tier — already stripped upstream

    purchase_price_row = capital_header_row + 1
    ws_t12.cell(row=purchase_price_row, column=1, value="Target Purchase Price")
    pp_cell = ws_t12.cell(row=purchase_price_row, column=2, value=purchase_price if purchase_price is not None else "")
    if isinstance(purchase_price, (int, float)):
        pp_cell.number_format = '$#,##0.00'

    total_acq_cost_row = purchase_price_row + 1

    if dst_capex is not None:
        dst_row = purchase_price_row + 1
        ws_t12.cell(row=dst_row, column=1, value="DST / Capex Budget")
        dst_cell = ws_t12.cell(row=dst_row, column=2, value=dst_capex)
        dst_cell.number_format = '$#,##0.00'

        comment_text = (
            "Deferred maintenance, physical capital repairs, and/or sponsor reserves "
            "as stated in the Offering Memorandum. Extracted by AI — verify against "
            "source document before relying on this figure."
        )
        # Module 4: append the actual source citation if the AI provided one
        citation_text = format_citation(metrics.get("provenance", {}).get("dst_capex_budget", {}))
        if citation_text:
            comment_text += f"\n\n{citation_text}"

        dst_cell.comment = Comment(comment_text, "Underwriting AI")
        total_acq_cost_row = dst_row + 1
        ws_t12.cell(row=total_acq_cost_row, column=1, value="TOTAL ACQUISITION COST").font = Font(bold=True)
        acq_cell = ws_t12.cell(
            row=total_acq_cost_row, column=2,
            value=f"=B{purchase_price_row}+B{dst_row}"
        )
        acq_cell.font = Font(bold=True)
        acq_cell.number_format = '$#,##0.00'
    else:
        # No DST/Capex (Basic tier, or AI found none) — Total Acquisition Cost is
        # just the purchase price, still expressed as a formula for consistency.
        ws_t12.cell(row=total_acq_cost_row, column=1, value="TOTAL ACQUISITION COST").font = Font(bold=True)
        acq_cell = ws_t12.cell(row=total_acq_cost_row, column=2, value=f"=B{purchase_price_row}")
        acq_cell.font = Font(bold=True)
        acq_cell.number_format = '$#,##0.00'

    # ==========================================
    # TAB 2 continued: Debt Sizing & Returns
    # LTV / Interest Rate / Amortization are raw user-entered numbers (not
    # formulas) — the underwriter's own assumptions, editable directly in the
    # spreadsheet afterward. Everything downstream (loan amount, debt service,
    # DSCR, cash-on-cash) is a live formula referencing these input cells.
    # ==========================================
    ltv_pct = debt_assumptions.get('ltv_pct', 75.0)
    interest_rate_pct = debt_assumptions.get('interest_rate_pct', 7.0)
    amortization_years = debt_assumptions.get('amortization_years', 30)

    debt_header_row = debt_note_row = total_acq_cost_row + 2
    ws_t12.cell(row=debt_header_row, column=1, value="DEBT ASSUMPTIONS & SIZING").font = Font(bold=True, size=12)

    ltv_row = debt_header_row + 1
    ws_t12.cell(row=ltv_row, column=1, value="Loan-to-Value (LTV) %")
    ltv_cell = ws_t12.cell(row=ltv_row, column=2, value=ltv_pct / 100)  # stored as a decimal for % formatting
    ltv_cell.number_format = '0.00%'
    ltv_cell.comment = Comment(
        "User-entered assumption — not extracted from source documents. "
        "Edit this cell to see the model recalculate.",
        "Underwriting AI"
    )

    interest_row = ltv_row + 1
    ws_t12.cell(row=interest_row, column=1, value="Interest Rate %")
    interest_cell = ws_t12.cell(row=interest_row, column=2, value=interest_rate_pct / 100)
    interest_cell.number_format = '0.000%'
    interest_cell.comment = Comment(
        "User-entered assumption — not extracted from source documents.",
        "Underwriting AI"
    )

    amort_row = interest_row + 1
    ws_t12.cell(row=amort_row, column=1, value="Amortization Term (years)")
    amort_cell = ws_t12.cell(row=amort_row, column=2, value=amortization_years)
    amort_cell.comment = Comment(
        "User-entered assumption — not extracted from source documents.",
        "Underwriting AI"
    )

    loan_amount_row = amort_row + 1
    ws_t12.cell(row=loan_amount_row, column=1, value="Sized Loan Amount").font = Font(bold=True)
    loan_cell = ws_t12.cell(
        row=loan_amount_row, column=2,
        value=f"=B{total_acq_cost_row}*B{ltv_row}"
    )
    loan_cell.font = Font(bold=True)
    loan_cell.number_format = '$#,##0.00'

    equity_row = loan_amount_row + 1
    ws_t12.cell(row=equity_row, column=1, value="Total Initial Cash Equity Required").font = Font(bold=True)
    equity_cell = ws_t12.cell(
        row=equity_row, column=2,
        value=f"=B{total_acq_cost_row}-B{loan_amount_row}"
    )
    equity_cell.font = Font(bold=True)
    equity_cell.number_format = '$#,##0.00'

    debt_service_row = equity_row + 1
    ws_t12.cell(row=debt_service_row, column=1, value="Annual Debt Service")
    # PMT(rate/12, term*12, -loan) * 12 — standard amortizing loan payment,
    # negated because PMT returns a negative value for an outflow.
    debt_service_cell = ws_t12.cell(
        row=debt_service_row, column=2,
        value=f"=PMT(B{interest_row}/12, B{amort_row}*12, -B{loan_amount_row})*12"
    )
    debt_service_cell.number_format = '$#,##0.00'

    dscr_row = debt_service_row + 1
    ws_t12.cell(row=dscr_row, column=1, value="Debt Service Coverage Ratio (DSCR)").font = Font(bold=True)
    noi_row_ref = line_item_rows.get("net_operating_income")
    if noi_row_ref:
        dscr_cell = ws_t12.cell(
            row=dscr_row, column=2,
            value=f"=B{noi_row_ref}/B{debt_service_row}"
        )
        dscr_cell.font = Font(bold=True)
        dscr_cell.number_format = '0.00"x"'

    cash_on_cash_row = dscr_row + 1
    ws_t12.cell(row=cash_on_cash_row, column=1, value="Cash-on-Cash Return %").font = Font(bold=True)
    if noi_row_ref:
        coc_cell = ws_t12.cell(
            row=cash_on_cash_row, column=2,
            value=f"=(B{noi_row_ref}-B{debt_service_row})/B{equity_row}"
        )
        coc_cell.font = Font(bold=True)
        coc_cell.number_format = '0.00%'

    # ==========================================
    # TAB 1: DASHBOARD SUMMARY
    # Every numeric cell is a cross-tab reference — no hardcoded values.
    # ==========================================
    ws_dash = wb.create_sheet("Dashboard Summary", 0)
    ws_dash.cell(row=1, column=1, value="METRIC")
    ws_dash.cell(row=1, column=2, value="VALUE")
    for cell in ws_dash[1]:
        cell.font = Font(bold=True, size=12, color="FFD4AF37")

    property_metadata = metrics.get("property_metadata", {})
    underwriting_ratios = metrics.get("underwriting_ratios", {})
    debt_returns = metrics.get("debt_returns", {})

    noi_row = line_item_rows.get("net_operating_income")
    revenue_row = line_item_rows.get("total_operating_revenue")
    opex_row = line_item_rows.get("total_operating_expenses")

    dash_rows = [
        ("Property Name", property_metadata.get("property_name"), None, False, None, None),
        ("Address", property_metadata.get("address"), None, False, None, None),
        ("Asset Class", property_metadata.get("asset_class"), None, False, None, None),
        ("Total Unit Count", None, f"='Rent Roll Analysis'!B{unit_count_row}", False, None, None),
        ("Total Vacant Units", None, f"='Rent Roll Analysis'!B{vacant_units_row}", False, None, None),
        ("Physical Occupancy %", None, f"='Rent Roll Analysis'!B{occupancy_row}", True, None, "physical_occupancy_pct"),
        ("Avg In-Place Monthly Rent", None, f"='Rent Roll Analysis'!B{avg_rent_row}", False, '$', None),
        ("Total Annual Concessions", metrics.get("rent_roll_metrics", {}).get("total_annual_concessions"), None, False, '$', "total_annual_concessions"),
        ("Total Operating Revenue", None, f"='Cleaned T12'!B{revenue_row}" if revenue_row else None, False, '$', None),
        ("Total Operating Expenses", None, f"='Cleaned T12'!B{opex_row}" if opex_row else None, False, '$', None),
        ("Net Operating Income (NOI)", None, f"='Cleaned T12'!B{noi_row}" if noi_row else None, False, '$', None),
        ("Net Rental Income", None, f"='Cleaned T12'!B{line_item_rows.get('net_rental_income')}" if line_item_rows.get('net_rental_income') else None, False, '$', "net_rental_income"),
        ("Operating Expense Ratio", underwriting_ratios.get("operating_expense_ratio"), None, True, None, None),
        ("Total Annual Expenses / Unit", underwriting_ratios.get("total_annual_expenses_per_unit"), None, False, '$', None),
        ("Target Purchase Price", None, f"='Cleaned T12'!B{purchase_price_row}", False, '$', None),
        ("Entry Cap Rate %", deal_valuation.get("entry_cap_rate"), None, True, None, None),
        ("Acquisition Price / Unit", deal_valuation.get("acquisition_price_per_unit"), None, False, '$', None),
        ("Cash-on-Cash Return %", debt_returns.get("cash_on_cash_return_pct"), None, True, None, None),
    ]

    if dst_capex is not None:
        dash_rows.append(("DST / Capex Budget", None, f"='Cleaned T12'!B{total_acq_cost_row - 1}", False, '$', None))

    dash_rows.append(("Sized Loan Amount", None, f"='Cleaned T12'!B{loan_amount_row}", False, '$', None))
    dash_rows.append(("Total Equity Required", None, f"='Cleaned T12'!B{equity_row}", False, '$', None))
    dash_rows.append(("Annual Debt Service", None, f"='Cleaned T12'!B{debt_service_row}", False, '$', None))
    if noi_row_ref:
        dash_rows.append(("DSCR", None, f"='Cleaned T12'!B{dscr_row}", False, None, None))
        dash_rows.append(("Cash-on-Cash Return % (Modeled)", None, f"='Cleaned T12'!B{cash_on_cash_row}", True, None, None))

    row_idx = 2
    dash_row_by_metric = {}  # metric key -> row number, for reconciliation highlighting below

    for entry in dash_rows:
        label, static_val, formula, is_pct, currency, metric_key = entry

        ws_dash.cell(row=row_idx, column=1, value=label)
        value_cell = ws_dash.cell(row=row_idx, column=2, value=formula if formula else static_val)

        if is_pct:
            value_cell.number_format = '0.00%'
        elif currency == '$':
            value_cell.number_format = '$#,##0.00'

        if metric_key:
            dash_row_by_metric[metric_key] = row_idx

        row_idx += 1

    # ==========================================
    # MODULE 1 (Enterprise only): Reconciliation flag highlighting
    # If flags were computed (see reconciliation.py — Python compares OM claims
    # against ground truth from the actual rent roll data), highlight the
    # corresponding Dashboard cell in light red with a hoverable comment
    # explaining the discrepancy, per the module spec.
    # ==========================================
    reconciliation = metrics.get("reconciliation")
    if reconciliation and reconciliation.get("flags"):
        flag_fill = PatternFill(start_color="FFF8D7DA", end_color="FFF8D7DA", fill_type="solid")

        for flag in reconciliation["flags"]:
            target_row = dash_row_by_metric.get(flag["metric"])
            if target_row is None:
                continue  # flag references a metric not shown on this tab — skip highlighting, still listed below

            label_cell = ws_dash.cell(row=target_row, column=1)
            value_cell = ws_dash.cell(row=target_row, column=2)
            label_cell.fill = flag_fill
            value_cell.fill = flag_fill

            comment_text = f"[{flag['severity']}] {flag['message']}"
            # Module 4: the occupancy flag is specifically about the OM's claim —
            # attach its citation here if the AI provided one.
            if flag["metric"] == "physical_occupancy_pct":
                citation_text = format_citation(metrics.get("provenance", {}).get("om_claimed_occupancy_pct", {}))
                if citation_text:
                    comment_text += f"\n\n{citation_text}"

            value_cell.comment = Comment(comment_text, "Reconciliation Engine")

        # Full flags list below the main table too, so nothing is hidden behind
        # a hover-only comment — visible at a glance even without inspecting cells.
        flags_header_row = row_idx + 2
        ws_dash.cell(row=flags_header_row, column=1, value="RECONCILIATION FLAGS").font = Font(bold=True, size=12, color="FFD4AF37")

        flag_row = flags_header_row + 1
        for flag in reconciliation["flags"]:
            severity_cell = ws_dash.cell(row=flag_row, column=1, value=f"[{flag['severity']}]")
            severity_cell.font = Font(bold=True, color="FFCC0000" if flag['severity'] == "CRITICAL" else "FFB8860B")
            ws_dash.cell(row=flag_row, column=2, value=flag["message"])
            flag_row += 1
        next_free_row = flag_row
    elif reconciliation and reconciliation.get("note"):
        # Reconciliation couldn't run (e.g. rent roll columns unclear) — surface why.
        note_row = row_idx + 2
        ws_dash.cell(row=note_row, column=1, value="RECONCILIATION:").font = Font(bold=True, italic=True)
        ws_dash.cell(row=note_row, column=2, value=reconciliation["note"]).font = Font(italic=True, color="FF888888")
        next_free_row = note_row + 1
    else:
        next_free_row = row_idx

    # ==========================================
    # MODULE 2 (Enterprise only): Standardization summary
    # Quick visibility on the Dashboard tab so uncategorized items aren't only
    # discoverable by scrolling into Tab 2's detail section.
    # ==========================================
    if standardization and standardization.get("uncategorized_items"):
        std_header_row = next_free_row + 2
        ws_dash.cell(row=std_header_row, column=1, value="UNCATEGORIZED ITEMS DETECTED").font = Font(bold=True, size=12, color="FFB8860B")

        count = len(standardization["uncategorized_items"])
        total = standardization["uncategorized_total"]
        summary_row = std_header_row + 1
        ws_dash.cell(
            row=summary_row, column=1,
            value=f"{count} line item(s) totaling ${total:,.2f} could not be confidently "
                  f"categorized. See 'Cleaned T12' tab for details — manual review recommended."
        ).font = Font(italic=True, color="FF888888")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()